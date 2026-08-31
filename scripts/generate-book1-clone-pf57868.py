#!/usr/bin/env python3
"""
Build Kenya UAT visual QA Excel EXACTLY like
  C:\\Users\\ThejanaD\\Downloads\\iPay Lite Testing (1).xlsx

Layout (Sheet1):
  Col A = Area (merged when multi-row)
  Col B = Issue / Concern text
  Col C = Screenshot (embedded, TwoCellAnchor) — NO header row
"""
from __future__ import annotations

import shutil
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor, OneCellAnchor, TwoCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from openpyxl.styles import Alignment, Border, Side, Font
from PIL import Image as PILImage

TEMPLATE = Path(r"C:\Users\ThejanaD\Downloads\iPay Lite Testing (1).xlsx")
ANN = Path(r"C:\Users\ThejanaD\QAFusionX\proof-2round-annotated-aug31")
RAW = Path(r"C:\Users\ThejanaD\QAFusionX\proof-2round-complete-aug31")
CHK = Path(r"C:\Users\ThejanaD\QAFusionX\proof-checker-password-aug31")
DONE = Path(r"C:\Users\ThejanaD\QAFusionX\proof-done-push-aug31")
THUMB = Path(r"C:\Users\ThejanaD\QAFusionX\proof-2round-annotated-aug31\_book1-thumbs")

# (Area, Issue, image candidates) — one row per issue like the iPay Lite (1) doc
ROWS: list[tuple[str, str, list[str]]] = [
    ("PF-58374 | PERC View",
     "Instalment-wise grace period fields are missing on the Penal Interest Rate Change View / Pending detail screen (PF-58496). Story remains PARTIAL — product bug, not QA process.",
     ["ANN-r2-74-view.png", "r2-74-view.png"]),
    ("PF-58374 | PERC Edit",
     "Go To Edit / detail path does not expose instalment-wise grace tier fields required by the AC (PF-58496 / PF-58497).",
     ["ANN-r2-74-edit.png", "r2-74-edit.png"]),
    ("PF-58374 | Template",
     "Penal interest template / settings screens show copy defects (e.g. INTREST) and no grace tier field (PF-58511).",
     ["ANN-r2-74-tpl-template.png", "r2-74-tpl-template.png"]),
    ("PF-58375 | Print Offer Letter",
     "Print Offer Letter UI is blank or stuck on Loading spinner; Joint/Business RTO templates are missing (PF-58500).",
     ["ANN-done-75-offer.png", "ANN-r2-75-offer.png", "r2-75-offer.png"]),
    ("PF-58375 | RTO Contract",
     "RTO contract / offer template path is incomplete for Joint and Business products (PF-58499).",
     ["ANN-r2-75-contract.png", "r2-75-contract.png"]),
    ("PF-58375 | RTO Inquiry",
     "Rent-to-Own inquiry returns incomplete content / No Data after search.",
     ["ANN-r2-75-inq-rto.png", "r2-75-inq-rto.png"]),
    ("PF-58376 | Schedule 404",
     "Deep link to schedule-monitory-dashboard returns HTTP 404 on Kenya UAT (PF-58418).",
     ["ANN-done-76-CASA-search.png", "76-CASA-search.png"]),
    ("PF-58376 | Schedule No Data",
     "After selecting module/dates, Schedule Monitory Dashboard search stays on No Data — Success+Error rows cannot be verified (PF-58418/58425/58426).",
     ["ANN-r2-76-search-CASA.png", "r2-76-search-CASA.png"]),
    ("PF-58376 | Schedule Process",
     "Process schedule screen remains empty / No Data.",
     ["ANN-r2-76-process.png", "r2-76-process.png"]),
    ("PF-58377 | Entity Creation",
     "Entity Creation page renders a blank / broken CSS shell with no usable form (PF-58512).",
     ["ANN-done-77-entity.png", "ANN-r2-77-entity-creation.png", "r2-77-entity-creation.png"]),
    ("PF-58377 | Supplier Creation",
     "Supplier Creation shows the same blank shell — main content does not load.",
     ["ANN-r2-77-supplier-creation.png", "r2-77-supplier-creation.png"]),
    ("PF-58377 | Pending duplicates",
     "Pending Entity Confirmation lists SUP0000002558 three times; Entity Name / NIC columns show dashes (PF-58513).",
     ["ANN-r2-77-pending-supplier-confirmation.png", "r2-77-pending-supplier-confirmation.png"]),
    ("PF-58378 | Value Date",
     "Non Counter Deposit View shows Value Date as '-' while Batch Date is populated (PF-58514).",
     ["ANN-r2-78-view.png", "r2-78-view.png"]),
    ("PF-58378 | NCD Create",
     "Create Non Counter Deposit is blocked by cash float / deposit-type validation on Save.",
     ["ANN-r2-78-save.png", "r2-78-save.png"]),
    ("PF-58380 | Receipt Reversal",
     "Receipt reversal screen is blank / empty — migrated receipt reverse cannot be verified.",
     ["ANN-r2-80-rev.png", "r2-80-rev.png"]),
    ("PF-58380 | Inquiry",
     "Loan/account inquiry searches (seed accounts) return no hit on Kenya UAT this cycle.",
     ["ANN-r2-80-inq-004225.png", "r2-80-inq-004225.png"]),
    ("PF-58380 | Reallocation",
     "Receipt reallocation and maintenance screens open empty with no actionable rows.",
     ["ANN-r2-80-realloc.png", "r2-80-realloc.png"]),
    ("PF-58383 | GBAF selector",
     "Deep Account Management routes are trapped on Select Banking & Finance Type (GBAF/IBAF) and never reach the feature screen (PF-58398/58416).",
     ["ANN-r2-83-account-management-manage-account.png", "r2-83-account-management-manage-account.png"]),
    ("PF-58383 | Ownership transfer",
     "Ownership transfer / owner-transfer-history remains unreachable past the Banking Type selector.",
     ["ANN-r2-83-maintenance-ownership-transfer.png", "r2-83-maintenance-ownership-transfer.png"]),
    ("PF-58383 | Account inquiry",
     "Account inquiry deep link is blocked by the same GBAF/IBAF selector trap.",
     ["ANN-r2-83-cNwNb-account-inquiry.png", "r2-83-cNwNb-account-inquiry.png"]),
    ("PF-58560 | Checker login",
     "RESOLVED: Checker MethmiB@lolctech.com can sign in via Use my password. Bug moved to In UAT.",
     ["ANN-checker-chk-final.png", "chk-final.png"]),
    ("PF-58560 | Checker PERC",
     "As MethmiB, Penal Interest Rate Change opens; Requests tab showed No Data on the approve attempt.",
     ["ANN-checker-chk-approve-attempt.png", "chk-approve-attempt.png"]),
    ("PF-58379 | Accrued Interest",
     "DONE N/A Kenya — Accrued Interest feature is not present on the cNwNb Kenya UAT build.",
     ["ANN-r2-na-79.png", "r2-na-79.png"]),
    ("PF-58381 | Document Request",
     "DONE N/A Kenya — Document Request workflow is not on the Kenya COB/Lending build.",
     ["ANN-r2-na-81.png", "r2-na-81.png"]),
    ("PF-58382 | Profit Sharing",
     "DONE N/A Kenya — Islamic profit-sharing ratio template is not deployed on Kenya UAT.",
     ["ANN-r2-na-82.png", "r2-na-82.png"]),
    ("PF-58384 | BRWNS SMS",
     "DONE N/A Kenya — BRWNS / SMS template management is not available under Common Settings.",
     ["ANN-r2-na-84.png", "r2-na-84.png"]),
]


def find_img(names: list[str]) -> Path | None:
    for n in names:
        for root in (ANN, DONE, CHK, RAW):
            p = root / n
            if p.exists():
                return p
    return None


def make_thumb(src: Path, dest: Path, max_w: int = 220) -> tuple[int, int]:
    im = PILImage.open(src).convert("RGB")
    w, h = im.size
    if w > max_w:
        nh = max(1, int(h * (max_w / w)))
        im = im.resize((max_w, nh), PILImage.Resampling.LANCZOS)
    else:
        max_w, nh = w, h
    dest.parent.mkdir(parents=True, exist_ok=True)
    im.save(dest, format="PNG", optimize=True)
    return max_w, nh


def add_image_col_c(ws, row_1based: int, img_path: Path, px_w: int, px_h: int) -> None:
    """Place image in column C (index 2) on the given 1-based row, TwoCellAnchor like template."""
    img = XLImage(str(img_path))
    img.width = px_w
    img.height = px_h
    # openpyxl rows/cols are 0-based in AnchorMarker
    # col 2 = C
    marker = AnchorMarker(col=2, colOff=pixels_to_EMU(4), row=row_1based - 1, rowOff=pixels_to_EMU(2))
    ext = XDRPositiveSize2D(pixels_to_EMU(px_w), pixels_to_EMU(px_h))
    # OneCellAnchor is enough and reliable
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor as OCA
    from openpyxl.drawing.xdr import XDRPoint2D

    img.anchor = OCA(_from=marker, ext=ext)
    ws.add_image(img)


def main() -> None:
    if THUMB.exists():
        shutil.rmtree(THUMB, ignore_errors=True)
    THUMB.mkdir(parents=True, exist_ok=True)

    # Clone column layout from template
    tw = load_workbook(TEMPLATE)
    tsheet = tw["Sheet1"]
    col_a_w = tsheet.column_dimensions["A"].width or 28.5
    col_b_w = tsheet.column_dimensions["B"].width or 133.0
    col_c_w = tsheet.column_dimensions["C"].width or 34.5
    tw.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    area_font = Font(bold=True, size=11)

    ws.column_dimensions["A"].width = col_a_w
    ws.column_dimensions["B"].width = col_b_w
    ws.column_dimensions["C"].width = max(col_c_w, 40)

    # No header row — same as iPay Lite Testing (1).xlsx (data starts row 2 in that file;
    # we start at row 1 for clarity but match 3-col visual shape)
    # Actually template starts content at row 2 with empty row 1 — match that.
    start = 2
    placed = 0
    missing: list[str] = []

    for i, (area, issue, names) in enumerate(ROWS):
        r = start + i
        ws.row_dimensions[r].height = 120
        c_a = ws.cell(r, 1, area)
        c_a.alignment = wrap
        c_a.font = area_font
        c_a.border = thin
        c_b = ws.cell(r, 2, issue)
        c_b.alignment = wrap
        c_b.border = thin
        c_c = ws.cell(r, 3, "")
        c_c.border = thin

        src = find_img(names)
        if not src:
            ws.cell(r, 3, f"(missing {names[0]})")
            missing.append(names[0])
            continue
        thumb = THUMB / f"r{r:02d}.png"
        pw, ph = make_thumb(src, thumb, max_w=240)
        # cap height for row
        if ph > 110:
            scale = 110 / ph
            pw = max(1, int(pw * scale))
            ph = 110
            # rewrite scaled
            im = PILImage.open(src).convert("RGB")
            im = im.resize((pw, ph), PILImage.Resampling.LANCZOS)
            im.save(thumb, format="PNG", optimize=True)
        add_image_col_c(ws, r, thumb, pw, ph)
        placed += 1

    outs = [
        Path(r"C:\Users\ThejanaD\Downloads\iPay Lite Kenya UAT PF-57868.xlsx"),
        Path(r"C:\Users\ThejanaD\Downloads\PF-57868-same-as-iPay-Lite-Testing.xlsx"),
        Path(r"E:\QAFusionX\workspaces\PF-57868\reports\iPay Lite Kenya UAT PF-57868.xlsx"),
        Path(r"E:\QAFusionX\workspaces\PF-57868\artifacts\iPay Lite Kenya UAT PF-57868.xlsx"),
    ]

    # Save once then copy (images + openpyxl double-save issue)
    primary = outs[0]
    # if locked, use alternate name
    try:
        if primary.exists():
            primary.unlink()
    except PermissionError:
        primary = Path(r"C:\Users\ThejanaD\Downloads\iPay Lite Kenya UAT PF-57868-NEW.xlsx")
        outs[0] = primary

    wb.save(primary)
    print("saved", primary, "mb", round(primary.stat().st_size / 1e6, 2), "images", placed)

    for p in outs[1:]:
        p.parent.mkdir(parents=True, exist_ok=True)
        try:
            shutil.copy2(primary, p)
            print("copied", p)
        except PermissionError:
            alt = p.with_name(p.stem + "-NEW" + p.suffix)
            shutil.copy2(primary, alt)
            print("LOCKED ->", alt)

    # verify
    v = load_workbook(primary)
    print("verify sheet", v.sheetnames, "imgs", len(v["Sheet1"]._images), "rows", v["Sheet1"].max_row)
    v.close()
    print("missing", missing)


if __name__ == "__main__":
    main()
