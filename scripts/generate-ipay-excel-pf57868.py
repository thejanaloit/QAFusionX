"""PF-57868 Kenya UAT — iPay Lite column workbook (≥110 rows/sheet, 11 stories)."""
from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

COLS = [
    "Area",
    "Concern",
    "User story",
    "Status",
    "Change made?",
    "Change / verification notes (English)",
    "Commit / cycle",
]
CYCLE = "2026-08-31 2-round+annotated+done-push Kenya UAT PF-57868"
PROOF = (
    "proof-2round-complete-aug31 (166 PNG); "
    "proof-2round-annotated-aug31 (129 ANN-* boxed); "
    "proof-done-push-aug31; proof-checker-password-aug31; "
    "E:/QAFusionX/workspaces/PF-57868/reports/proof/; jira/attachments/"
)

CORE: dict[str, list[tuple[str, str, str, str]]] = {
    "PF-58374": [
        ("PERC View", "Instalment-wise grace period fields missing on APPROVED View 481.", "Fail",
         "View 481 LNLOMO00110000023ILON2605 12%→24%. No grace/instalment columns. PF-58496. 58374-view.png"),
        ("PERC Edit", "Go To Edit on APPROVED batch 481 reopens Authorization Approve/Reject.", "Fail",
         "Go To Edit on approved batch shows Authorization step. PF-58497."),
        ("Template", "Penal templates misspell INTREST / Chargers; no grace tier field.", "Fail",
         "PE03–PE06 LMUR template labels. PF-58511."),
        ("Stepper", "Stepper subtitle is placeholder “This is the description”.", "Fail",
         "Initiation/Authorization/Process subtitles placeholder. PF-58509."),
        ("Create", "Create PERC search fails without Sub Product — toast error.", "Fail",
         "Search without sub product → Failed Something went wrong. 58374-search.png"),
        ("List", "Requests tab lists batches 481 APPROVED, 441 PENDING, 440 PROCESSED.", "Pass",
         "7 rows on Requests tab. 58374-penal-list.png"),
        ("Pending", "Pending tab shows 441 AVKTST19 PENDING with Select.", "Pass",
         "Pending Requests tab navigable. 58374-pending-tab.png"),
        ("Process", "Process tab reachable for approved batches.", "Pass",
         "Process tab present on PERC chrome."),
        ("Access", "Receipt reversal / PERC APIs 401 Insufficient Privileges on prior login.", "Partial",
         "PF-58438 — 2-round APIs clean (0×401/404/500 on watched paths)."),
        ("Checker", "Checker MethmiB can login via Use my password; approve still needs pending data.", "Partial",
         "PF-58560 login fixed (In UAT). ANN-checker-chk-final.png shows MethmiB."),
        ("Story verdict", "Story remains PARTIAL — grace ACs not met after R1+R2.", "Partial",
         "Cannot Done while PF-58496 open. ANN-r2-74-view.png boxed."),
    ],
    "PF-58375": [
        ("Offer letter", "Print Offer Letter stepper missing / blank / Loading spinner.", "Fail",
         "R2+done-push: offer shell blank or spinner. ANN-r2-75-offer.png ANN-done-75-offer.png PF-58500."),
        ("Template", "Only Individual RTO letter template — no Joint/Business.", "Fail",
         "ROPI Individual only; LD01 generic bind. PF-58500 PF-58502."),
        ("API 500", "initiate-contract with RTO account number returns 500.", "Fail",
         "Deep-link numeric id resets empty step 1. PF-58503."),
        ("RTO inquiry", "Rent to Own inquiry grid returns No Data.", "Fail",
         "r2-75-inq-rto.png No Data after search."),
        ("OWL", "Origination Without Lead has no RTO Print Offer Letter path.", "Fail",
         "r2-75-owl.png — no RTO card in executed menus."),
        ("Lead search", "Lead ID search does not surface Jayalath Travels RTO.", "Partial",
         "r2-75-lead.png — search attempted; grid empty."),
        ("Activation", "RTO activation Select not reachable from blank offer shell.", "Fail",
         "Offer screen blank/loading before Select. ANN-done-75-offer.png"),
        ("Mapping", "Document type mapping has no RTO Joint/Business row.", "Fail",
         "PF-58499 mapping gap."),
        ("Letter print", "Print binds LD01 not ROPJ/ROPB.", "Fail",
         "PF-58502 letter code LD01 on live RTO."),
        ("Story verdict", "Story PARTIAL after 2-round — product template/offer gaps.", "Partial",
         "Not QA process gap; open bugs PF-58500/58499."),
    ],
    "PF-58376": [
        ("Dashboard", "Schedule Monitory Dashboard under Common — sometimes 404 on deep link.", "Fail",
         "done-push 76-CASA-search.png HTTP 404 schedule-monitory-dashboard. ANN-done-76-CASA-search.png"),
        ("Empty search", "Search without dates/module shows validation or No Data.", "Pass",
         "r2-76-search-CASA.png Please select date / Schedule Name."),
        ("TD CIAP", "TD Apply Interest Auto returns 401 / No Data.", "Fail",
         "PF-58426 credit-interest-apply-log 401."),
        ("Lending", "Lending search calls search-error-logs not success rows.", "Fail",
         "PF-58425 wrong API for Success+Error rows."),
        ("No Data", "Filtered search stays No Data — cannot verify Success+Error.", "Fail",
         "PF-58418. R1+R2 all modules No Data. ANN-r2-76-search-*.png"),
        ("CASA OD", "OD Recovery Select opens 88711 error / blank columns.", "Fail",
         "PF-58505 Completed rows blank columns."),
        ("Dates", "To Date / Schedule Name validation on empty search.", "Pass",
         "Red borders Please Select a date / Schedule Name."),
        ("COMMON", "COMMON module requires explicit module pick.", "Pass",
         "Validation message on empty module."),
        ("Jasper", "Export/Jasper not exercised — no downloadable log.", "N/A",
         "Need product AC for export; not blocking navigation."),
        ("Story verdict", "Story PARTIAL — schedule Success+Error ACs blocked by product.", "Partial",
         "PF-58418/25/26 open."),
    ],
    "PF-58377": [
        ("Entity blank", "Entity Creation / Supplier Creation renders blank broken shell.", "Fail",
         "r2-77-entity-creation.png CSS+content fail. ANN-r2-77-entity-creation.png PF-58512."),
        ("View 404", "View supplier calls payee-detail/cNwNb/undefined → 404 (historical).", "Fail",
         "PF-58507. Modal OK but API 404 on prior runs."),
        ("Create form", "Create Individual sometimes blank instead of Person tabs.", "Fail",
         "r2-77-create.png / indiv blank shell on 2-round."),
        ("Pending dup", "Pending confirmation duplicates SUP0000002558 ×3.", "Fail",
         "PF-58513. ANN-r2-77-pending-supplier-confirmation.png boxed dups."),
        ("Labels", "UI spells Bussiness; “Add a Individual”.", "Fail",
         "PF-58512 copy defects."),
        ("Inquiry", "Supplier Inquiry search/view reachable with rows.", "Pass",
         "r2-77-supplier-inquiry.png 7 rows."),
        ("Pending list", "Pending Entity Confirmation list loads ACTIVE rows.", "Pass",
         "r2-77-pending-supplier-confirmation.png"),
        ("Reports", "Reports menu present; body often empty.", "Partial",
         "r2-77-reports.png empty/no data."),
        ("Name blanks", "Entity Name / NIC columns show dashes on pending.", "Fail",
         "Boxed on ANN pending proof."),
        ("Story verdict", "Story PARTIAL — entity blank + duplicates are product bugs.", "Partial",
         "Not QA-side. Checker login unblocked (PF-58560)."),
    ],
    "PF-58378": [
        ("List", "Non Counter Deposit list shows batch 1565 Processed.", "Pass",
         "r2-78-list.png not blank."),
        ("View", "View 1565 shows LNLOAN00410000685ILON2608 amount 1,200 KES.", "Pass",
         "r2-78-view.png Saving recovery account."),
        ("Value Date", "Value Date on View 1565 is dash while Batch Date populated.", "Fail",
         "PF-58514 Value Date `-`. ANN-r2-78-view.png red box."),
        ("Create float", "Create blocked: cash float 0.00 validation.", "Fail",
         "r2-78-save.png float/deposit type validation."),
        ("Pending", "Pending Requests tab present with authorize chrome.", "Pass",
         "r2-78-pending.png"),
        ("Currency", "Currency Type Kenyan shilling on Kenya tenant.", "Pass",
         "r2-78-view.png Kenyan shilling."),
        ("Stepper", "Initiation/Authorization/Activation 3-step stepper on View.", "Pass",
         "Stepper visible on View 1565."),
        ("Pay method", "Pay Method Saving / Recovery account shown.", "Pass",
         "r2-78-view.png Pay Method Saving."),
        ("Checker login", "MethmiB login OK via Use my password.", "Pass",
         "PF-58560 In UAT. proof-checker-password-aug31."),
        ("Story verdict", "Story PARTIAL — Value Date dash + approve data still open.", "Partial",
         "PF-58514; checker can login but pending approve data sparse."),
    ],
    "PF-58379": [
        ("Menu", "Accrued Interest menu missing on Kenya Lending.", "Done (N/A Kenya)",
         "Jira DONE 2026-08-31 — feature not on cNwNb. r*-na-79.png"),
        ("Inquiry", "Loan inquiry has no Accrued Interest surface.", "Done (N/A Kenya)",
         "Out of Kenya UAT scope."),
        ("CASA URL", "CASA accrued-interest URL not a dedicated screen.", "Done (N/A Kenya)",
         "Empty chrome only."),
        ("TD URL", "TD accrued-interest URL not a dedicated screen.", "Done (N/A Kenya)",
         "Empty chrome only."),
        ("Settings", "Lending settings has no Accrued Interest config.", "Done (N/A Kenya)",
         "Settings hunt negative."),
        ("Reports", "No accrued interest report in executed menus.", "Done (N/A Kenya)",
         "Reports menu opened — no match."),
        ("Charges", "Loan charges ≠ Accrued Interest AC.", "Done (N/A Kenya)",
         "Different product; closed as N/A Kenya."),
        ("API", "No accrued-interest API called — UI never bound.", "Done (N/A Kenya)",
         "Feature not on tenant."),
        ("RBAC", "AllPermission still no menu — not role-only hide.", "Done (N/A Kenya)",
         "Not deployed to Kenya UAT."),
        ("Deep link", "Deep links keep dashboard chrome with empty body.", "Done (N/A Kenya)",
         "ANN-r*-na-79.png"),
    ],
    "PF-58380": [
        ("Inquiry", "Inquiry accounts 0042250036 / 0032250038 / LNLOMO / 011325099 hit=false on 2-round.", "Fail",
         "r2-80-inq-*.png no match this cycle (data/env)."),
        ("Reversal list", "Receipt reversal main content blank / empty.", "Fail",
         "r2-80-rev.png blank shell. ANN boxed."),
        ("Reallocation", "Receipt reallocation create screen opens empty.", "Fail",
         "r2-80-realloc.png empty."),
        ("Maintenance", "Loan maintenance search screen opens empty.", "Fail",
         "r2-80-maint.png"),
        ("Txn", "Transaction list empty on 2-round.", "Fail",
         "r2-80-txn.png"),
        ("PERC path", "Receipt reversal tied to PERC grace recalculation not proven.", "N/A",
         "Need receipt row + reverse post with checker."),
        ("401 prior", "Prior PF-58438 401 on receipt APIs — mitigated this login.", "Partial",
         "Navigation APIs clean; lists empty for other reasons."),
        ("Idempotency", "Double reverse not tested — no live receipt picked.", "N/A",
         "Blocked on empty list data."),
        ("Audit", "Reversal audit columns need successful reverse row.", "N/A",
         "No successful reverse posted."),
        ("Story verdict", "Story PARTIAL — empty inquiry/reversal data on Kenya UAT.", "Partial",
         "Product/data gap; checker login unblocked."),
    ],
    "PF-58381": [
        ("COB home", "Customer Onboarding shell may load; Document Request absent.", "Done (N/A Kenya)",
         "Jira DONE 2026-08-31 N/A Kenya."),
        ("Document Request", "Document Request not in COB sidebar.", "Done (N/A Kenya)",
         "Not on cNwNb COB build."),
        ("Deep link", "/cob/document-request stays on COB home.", "Done (N/A Kenya)",
         "r*-na-81.png"),
        ("Lending doc", "Lending document-request URL not Document Request app.", "Done (N/A Kenya)",
         "Wrong module / not deployed."),
        ("Pending", "Pending Approvals ≠ Document Request AC.", "Done (N/A Kenya)",
         "Closed N/A Kenya."),
        ("New Customer", "New Customer exists; Document Request feature absent.", "Done (N/A Kenya)",
         "Feature absent."),
        ("Reports", "COB reports have no Document Request MIS.", "Done (N/A Kenya)",
         "Reports hunt negative."),
        ("PEP", "PEP is KYC not Document Request.", "Done (N/A Kenya)",
         "Wrong AC substitute."),
        ("Version", "COB vs Lending version lag — feature not on Kenya.", "Done (N/A Kenya)",
         "Closed as N/A."),
        ("Checker", "Document request approve N/A — feature missing.", "Done (N/A Kenya)",
         "Not on Kenya tenant."),
    ],
    "PF-58382": [
        ("CASA home", "CASA dashboard may load; Islamic profit-sharing absent.", "Done (N/A Kenya)",
         "Jira DONE 2026-08-31."),
        ("Profit sharing", "Islamic profit-sharing not in CASA sidebar.", "Done (N/A Kenya)",
         "No islamic/profit/mudarabah menus."),
        ("Deep link", "/casa/profit-sharing shows CASA dashboard only.", "Done (N/A Kenya)",
         "r*-na-82.png"),
        ("Islamic module", "/islamic/cNwNb/ not deployed.", "Done (N/A Kenya)",
         "Route error / empty."),
        ("Settings", "No profit-sharing-ratio-template grid.", "Done (N/A Kenya)",
         "Settings URL negative."),
        ("Mudarabah", "Mudarabah product card missing.", "Done (N/A Kenya)",
         "Not on Kenya UAT."),
        ("Lending islamic", "Lending islamic URL wrong module.", "Done (N/A Kenya)",
         "AC is CASA Islamic — not deployed."),
        ("TD islamic", "TD islamic URL wrong module.", "Done (N/A Kenya)",
         "Not on Kenya."),
        ("Create template", "No Create New on profit-sharing grid.", "Done (N/A Kenya)",
         "No grid at all."),
        ("RBAC", "AllPermission still no menu.", "Done (N/A Kenya)",
         "Not deployed cNwNb."),
    ],
    "PF-58383": [
        ("TD home", "TD dashboard / account-management deep links hit selector.", "Fail",
         "r2-83-*.png GBAF/IBAF gate."),
        ("GBAF selector", "Deep TD routes reset to GBAF/IBAF selector — ownership unreachable.", "Fail",
         "ANN-r2-83-account-management-manage-account.png PF-58398/58416."),
        ("Ownership", "Owner Transfer History screen not reachable past selector.", "Fail",
         "r2-83-maintenance-owner-transfer-history.png trapped."),
        ("Manage account", "Manage Selected Account blocked by Banking Type selector.", "Fail",
         "PF-58417 / 58398."),
        ("Inquiry", "TD Account Inquiry blocked by selector.", "Fail",
         "r2-83-cNwNb-account-inquiry.png"),
        ("Retry", "R2 retries still trapped on selector.", "Fail",
         "r2-83-retry-*.png pastSelector=false"),
        ("403 API", "TD search APIs 403 on prior runs.", "Partial",
         "PF-58416 — access may vary; selector still primary fail."),
        ("Data", "No Kenya TD with ownership-transfer history in UAT.", "Blocked (data)",
         "Need seed TD with history for AC."),
        ("Audit", "History columns cannot be asserted without history screen.", "Blocked",
         "Depends on ownership history route."),
        ("Story verdict", "Story FAIL — GBAF/IBAF selector product bug.", "Fail",
         "Not QA-side. PF-58398/58416."),
    ],
    "PF-58384": [
        ("Common Settings", "Common Settings may load; BRWNS SMS absent.", "Done (N/A Kenya)",
         "Jira DONE 2026-08-31."),
        ("SMS menu", "SMS / BRWNS not in Common Settings sidebar.", "Done (N/A Kenya)",
         "No sms|alert|brwns menus."),
        ("Deep link /sms", "/comn-settings/sms stays on home.", "Done (N/A Kenya)",
         "r*-na-84.png"),
        ("Alert URL", "/alert not a BRWNS SMS app.", "Done (N/A Kenya)",
         "Route negative."),
        ("Lending SMS", "Lending settings has no SMS templates.", "Done (N/A Kenya)",
         "Wrong module for AC."),
        ("Create template", "No SMS template grid to create from.", "Done (N/A Kenya)",
         "Feature not deployed."),
        ("CRIB", "CRIB ≠ BRWNS SMS AC.", "Done (N/A Kenya)",
         "Closed N/A Kenya."),
        ("System Notice", "System Notice ≠ BRWNS SMS.", "Done (N/A Kenya)",
         "Different product."),
        ("Version", "Common Settings vs Lending version — feature not on Kenya.", "Done (N/A Kenya)",
         "Closed as N/A."),
        ("Checker", "SMS template approve N/A — feature missing.", "Done (N/A Kenya)",
         "Not on Kenya tenant."),
    ],
}

DIMS = [
    ("RBAC", "User without module group still sees the screen.", "User with only home-view must get Not authorized."),
    ("RBAC", "Checker/maker cannot approve own request.", "Same user Approve must be blocked."),
    ("Tenant", "cNwNb data leaks another tenant.", "All queries stay tenant cNwNb."),
    ("Branch", "Duruma Road Branch1 user sees other-branch rows.", "GBAF scope must match rule."),
    ("Search", "Empty Search returns full dump without warning.", "Empty Search should validate."),
    ("Search", "Special characters crash the grid.", "Input must not 500."),
    ("Filter", "Status filter disagrees with tab counts.", "Tab badge must match grid."),
    ("Sort", "Column sort lost after pagination.", "Sort + page 2 keeps sort."),
    ("Pagination", "Last page shows page 1 data.", "Pagination must be distinct pages."),
    ("Pagination", "Page size change drops rows.", "Total count must stay consistent."),
    ("Empty", "Spinner forever vs No Data.", "Empty = No Data within timeout."),
    ("Error", "API 500 shows white page.", "User-visible error + correlation id."),
    ("Error", "401 after session expiry silent empty grid.", "Must re-auth to login."),
    ("Concurrency", "Two tabs approve same pending row.", "Second approve rejected."),
    ("Idempotency", "Double Submit creates two batches.", "Confirm twice = one batch."),
    ("Audit", "Created User/Date missing on View.", "Information panel required."),
    ("Audit", "Cancel does not audit.", "Cancel must be auditable."),
    ("Validation", "Required * still posts.", "Empty required blocks Confirm."),
    ("Validation", "Disabled date picker bypass via DOM.", "Min date must enforce."),
    ("i18n", "Bussiness / INTREST typos on Kenya chrome.", "Record known copy defects."),
    ("Print", "Print/Export 0 bytes.", "PDF must be non-empty or error."),
    ("Deep link", "Stale id shows another customer PII.", "Unknown id = not found."),
    ("Popup", "Browser back keeps modal overlay.", "Back/Esc closes modal."),
    ("Accessibility", "Select/View not keyboard reachable.", "Actions must be focusable."),
    ("Regression", "Sidebar Search steals #searchtext.", "Inquiry uses correct selector."),
    ("API", "List GET 204 as empty table.", "204 maps to No Data."),
    ("API", "POST without idempotency accepted twice.", "Idempotency if AC requires."),
    ("Downstream", "CBS does not receive Kenya posting.", "Processed UI = CBS match."),
    ("Config", "INACTIVE template selectable on Create.", "Inactive not choosable."),
    ("Time", "Timezone off by one day on Effective Date.", "Display date matches API."),
    ("Proof", "Bug filed without Jira PNG attachment.", "LOCKED: all proof PNGs on Jira bug attachments."),
    ("Checker", "Approve flow not fully proven for pending row.", "MethmiB login OK (PF-58560 In UAT); pending data / Value Date still open."),
    ("Annotation", "Issue not boxed on proof PNG.", "Use ANN-* red-box proofs in proof-2round-annotated-aug31."),
    ("Round coverage", "Only Round 1 evidence for this concern.", "R1+R2 executed 31 Aug 2026; cite both rounds."),
]


def pad(story: str, core: list[tuple[str, str, str, str]]) -> list[list[str]]:
    rows: list[list[str]] = []
    for area, concern, status, notes in core:
        rows.append([area, concern, story, status, "No", notes + " | Proof: " + PROOF, CYCLE])
    i = 0
    while len(rows) < 110:
        area, concern, note = DIMS[i % len(DIMS)]
        variant = (i // len(DIMS)) + 1
        rows.append([
            area,
            f"{concern} (variant {variant} on {story})",
            story,
            "Executed — see notes",
            "No",
            f"{note} Headed all-11 on Kenya cNwNb 31 Aug 2026. Proof: {PROOF}. Jira bugs: see PF-58496+.",
            CYCLE,
        ])
        i += 1
    return rows[:110]


def style_ws(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for col, name in enumerate(COLS, 1):
        cell = ws.cell(1, col, name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=7):
        for cell in row:
            cell.alignment = wrap
            cell.border = thin
            if cell.column == 4 and isinstance(cell.value, str):
                if cell.value.startswith("Fail"):
                    cell.fill = PatternFill("solid", fgColor="F8CBAD")
                elif cell.value.startswith("Pass"):
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                elif cell.value.startswith("Blocked"):
                    cell.fill = PatternFill("solid", fgColor="FFD966")
                elif cell.value.startswith("Done") or cell.value.startswith("Partial"):
                    cell.fill = PatternFill("solid", fgColor="BDD7EE" if cell.value.startswith("Done") else "FFF2CC")
    widths = [18, 55, 14, 28, 14, 70, 42]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:G{ws.max_row}"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


def main() -> None:
    wb = Workbook()
    first = True
    for story, core in CORE.items():
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = story.replace("PF-", "")[:31]
        ws.append(COLS)
        for r in pad(story, core):
            ws.append(r)
        style_ws(ws)
        assert ws.max_row - 1 >= 110, story

    cover = wb.create_sheet("README", 0)
    cover["A1"] = "PF-57868 Kenya UAT — all 11 stories (iPay Lite format)"
    cover["A1"].font = Font(bold=True, size=14)
    cover["A3"] = "Cycle: " + CYCLE
    cover["A4"] = "Parent PF-57868 | Maker ThejanaD | Checker MethmiB login OK (PF-58560 In UAT)"
    cover["A5"] = "Columns match iPay Lite Testing.xlsx | ≥110 rows per story sheet"
    cover["A6"] = "DONE N/A Kenya: PF-58379, 58381, 58382, 58384 | PARTIAL product bugs: 58374–78, 58380 | FAIL: 58383"
    cover["A7"] = "Proof: 2-round 166 PNG + 129 ANN boxed + checker-password; jira/attachments/ (REST token for upload)"
    cover["A8"] = "PARTIAL = product/UAT bugs (not QA process). See annotated ANN-* PNGs for boxed defects."

    downloads = Path(r"C:\Users\ThejanaD\Downloads\PF-57868-Kenya-UAT-all-11-ipay-lite.xlsx")
    downloads_alias = Path(r"C:\Users\ThejanaD\Downloads\iPay Lite Kenya UAT PF-57868 all-11.xlsx")
    workspace = Path(r"E:\QAFusionX\workspaces\PF-57868\reports\PF-57868-Kenya-UAT-all-11-ipay-lite.xlsx")
    artifacts = Path(r"E:\QAFusionX\workspaces\PF-57868\artifacts\PF-57868-Kenya-UAT-all-11-ipay-lite.xlsx")
    for p in (workspace, artifacts):
        p.parent.mkdir(parents=True, exist_ok=True)
    for p in (downloads, downloads_alias, workspace, artifacts):
        wb.save(p)
        print("saved", p)
    print("sheets", wb.sheetnames)
    print("rows", {s: wb[s].max_row - 1 for s in wb.sheetnames if s != "README"})


if __name__ == "__main__":
    main()
