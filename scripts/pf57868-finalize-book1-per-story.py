#!/usr/bin/env python3
"""
PF-57868 finalize pack:
1) Fine-tuned red-box annotations (tight issue regions)
2) Book1-style Excel (Area | Issue | Screenshot) — ONE file per story
3) Output directory: reports/book1-per-story/ + Downloads mirror
"""
from __future__ import annotations

import json
import shutil
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image, ImageDraw, ImageFont

SRC = Path(r"C:\Users\ThejanaD\QAFusionX\proof-2round-complete-aug31")
DONE = Path(r"C:\Users\ThejanaD\QAFusionX\proof-done-push-aug31")
CHK = Path(r"C:\Users\ThejanaD\QAFusionX\proof-checker-password-aug31")
ANN_OUT = Path(r"C:\Users\ThejanaD\QAFusionX\proof-fine-tuned-aug31")
THUMB = ANN_OUT / "_thumbs"

OUT_DIR = Path(r"E:\QAFusionX\workspaces\PF-57868\reports\book1-per-story")
DL_DIR = Path(r"C:\Users\ThejanaD\Downloads\PF-57868-book1-per-story")
ART_DIR = Path(r"E:\QAFusionX\workspaces\PF-57868\artifacts\book1-per-story")

W, H = 1520, 960

try:
    FONT = ImageFont.truetype("arialbd.ttf", 15)
    FONT_SM = ImageFont.truetype("arial.ttf", 13)
except Exception:
    FONT = ImageFont.load_default()
    FONT_SM = FONT


def R(x1, y1, x2, y2):
    """fractions → pixels"""
    return (int(x1 * W), int(y1 * H), int(x2 * W), int(y2 * H))


# Fine-tuned boxes: tight on the actual defect (not whole page)
# key = source png name (search SRC/DONE/CHK)
FINE: dict[str, list[tuple[tuple[int, int, int, int], str]]] = {
    # 58374 — pending list mistaken for view / missing grace on detail chrome
    "r2-74-view.png": [
        (R(0.22, 0.30, 0.98, 0.42), "ISSUE: Pending list — no instalment-wise grace fields (PF-58496)"),
        (R(0.55, 0.48, 0.72, 0.78), "PENDING badges only — AC grace UI absent"),
    ],
    "r2-74-edit.png": [
        (R(0.22, 0.28, 0.72, 0.55), "ISSUE: Detail/edit missing grace period controls (PF-58496)"),
    ],
    "r2-74-tpl-template.png": [
        (R(0.28, 0.22, 0.85, 0.55), "ISSUE: Template spelling / missing grace tier (PF-58511)"),
    ],
    # 58375
    "r2-75-offer.png": [
        (R(0.22, 0.14, 0.98, 0.88), "ISSUE: Print Offer Letter blank — stepper missing (PF-58500)"),
    ],
    "75-offer.png": [
        (R(0.40, 0.35, 0.70, 0.65), "ISSUE: Stuck Loading spinner — Offer UI never renders (PF-58500)"),
        (R(0.78, 0.02, 0.92, 0.08), "Header still Loading…"),
    ],
    "r2-75-contract.png": [
        (R(0.22, 0.14, 0.98, 0.88), "ISSUE: Contract/template path incomplete (PF-58499)"),
    ],
    "r2-75-inq-rto.png": [
        (R(0.30, 0.40, 0.70, 0.70), "ISSUE: RTO inquiry No Data / empty"),
    ],
    # 58376
    "76-CASA-search.png": [
        (R(0.28, 0.28, 0.72, 0.62), "ISSUE: HTTP 404 — schedule-monitory-dashboard (PF-58418)"),
    ],
    "r2-76-search-CASA.png": [
        (R(0.42, 0.28, 0.62, 0.40), "ISSUE: To Date validation — Please Select a date (PF-58425)"),
        (R(0.42, 0.40, 0.68, 0.50), "ISSUE: Schedule Name required empty"),
        (R(0.35, 0.55, 0.65, 0.78), "ISSUE: No Data — Success+Error not verifiable (PF-58418)"),
    ],
    "r2-76-process.png": [
        (R(0.30, 0.45, 0.70, 0.75), "ISSUE: Process schedule empty / No Data"),
    ],
    # 58377
    "r2-77-entity-creation.png": [
        (R(0.35, 0.08, 0.98, 0.88), "ISSUE: Main content blank — Entity Creation failed to render (PF-58512)"),
        (R(0.02, 0.05, 0.28, 0.55), "Shell only: unstyled menu, no form"),
    ],
    "r2-77-supplier-creation.png": [
        (R(0.35, 0.08, 0.98, 0.88), "ISSUE: Supplier Creation blank shell (PF-58512)"),
    ],
    "r2-77-pending-supplier-confirmation.png": [
        (R(0.22, 0.42, 0.98, 0.62), "ISSUE: Duplicate SUP0000002558 ×3 (PF-58513)"),
        (R(0.36, 0.42, 0.58, 0.78), "ISSUE: Entity Name / NIC show '-'"),
    ],
    # 58378 — tight on Value Date row only
    "r2-78-view.png": [
        (R(0.34, 0.30, 0.52, 0.38), "ISSUE: Value Date = '-' (PF-58514)"),
    ],
    "r2-78-save.png": [
        (R(0.30, 0.25, 0.75, 0.55), "ISSUE: Create blocked — float / deposit validation"),
    ],
    # 58380
    "r2-80-rev.png": [
        (R(0.22, 0.12, 0.98, 0.88), "ISSUE: Receipt reversal blank — no reverse rows"),
    ],
    "r2-80-inq-004225.png": [
        (R(0.30, 0.35, 0.70, 0.70), "ISSUE: Inquiry hit=false for seed account"),
    ],
    "r2-80-realloc.png": [
        (R(0.22, 0.12, 0.98, 0.88), "ISSUE: Reallocation empty"),
    ],
    # 58383 — tight on GBAF modal cards
    "r2-83-account-management-manage-account.png": [
        (R(0.24, 0.20, 0.76, 0.72), "ISSUE: GBAF/IBAF selector traps deep route (PF-58398/58416)"),
        (R(0.50, 0.42, 0.74, 0.68), "GBAF card — Kenya should proceed past this gate"),
    ],
    "r2-83-maintenance-ownership-transfer.png": [
        (R(0.24, 0.20, 0.76, 0.72), "ISSUE: Ownership transfer blocked by selector"),
    ],
    "r2-83-cNwNb-account-inquiry.png": [
        (R(0.24, 0.20, 0.76, 0.72), "ISSUE: Account inquiry blocked by selector"),
    ],
    # checker
    "chk-final.png": [
        (R(0.78, 0.02, 0.99, 0.10), "OK: MethmiB@lolctech.com logged in (PF-58560)"),
    ],
    "chk-approve-attempt.png": [
        (R(0.78, 0.02, 0.99, 0.10), "OK: Checker MethmiB on PERC"),
        (R(0.40, 0.48, 0.70, 0.72), "ISSUE: Requests No Data on approve attempt"),
    ],
    # N/A
    "r2-na-79.png": [(R(0.25, 0.20, 0.90, 0.80), "N/A Kenya: Accrued Interest not on build")],
    "r2-na-81.png": [(R(0.25, 0.20, 0.90, 0.80), "N/A Kenya: Document Request not on build")],
    "r2-na-82.png": [(R(0.25, 0.20, 0.90, 0.80), "N/A Kenya: Profit Sharing not on build")],
    "r2-na-84.png": [(R(0.25, 0.20, 0.90, 0.80), "N/A Kenya: BRWNS SMS not on build")],
}

# story → rows for Book1 excel (area, issue, source png for fine annotate)
STORY_ROWS: dict[str, list[tuple[str, str, str]]] = {
    "PF-58374": [
        ("PERC View", "Instalment-wise grace period fields missing on View/Pending detail (PF-58496).", "r2-74-view.png"),
        ("PERC Edit", "Edit/detail missing grace period controls (PF-58496/58497).", "r2-74-edit.png"),
        ("Template", "Template spelling / missing grace tier (PF-58511).", "r2-74-tpl-template.png"),
    ],
    "PF-58375": [
        ("Print Offer Letter", "Offer UI blank or Loading spinner; Joint/Business templates missing (PF-58500).", "r2-75-offer.png"),
        ("Offer Loading", "Done-push: stuck Loading spinner — Offer never renders (PF-58500).", "75-offer.png"),
        ("RTO Contract", "Contract/template path incomplete (PF-58499).", "r2-75-contract.png"),
        ("RTO Inquiry", "RTO inquiry No Data / empty.", "r2-75-inq-rto.png"),
    ],
    "PF-58376": [
        ("Schedule 404", "schedule-monitory-dashboard HTTP 404 (PF-58418).", "76-CASA-search.png"),
        ("Schedule No Data", "Search stays No Data — Success+Error not verifiable (PF-58418/25/26).", "r2-76-search-CASA.png"),
        ("Schedule Process", "Process schedule empty / No Data.", "r2-76-process.png"),
    ],
    "PF-58377": [
        ("Entity Creation", "Entity Creation blank / broken shell (PF-58512).", "r2-77-entity-creation.png"),
        ("Supplier Creation", "Supplier Creation blank shell (PF-58512).", "r2-77-supplier-creation.png"),
        ("Pending duplicates", "SUP0000002558 ×3; Name/NIC dashes (PF-58513).", "r2-77-pending-supplier-confirmation.png"),
    ],
    "PF-58378": [
        ("Value Date", "NCD View Value Date shows '-' (PF-58514).", "r2-78-view.png"),
        ("NCD Create", "Create blocked by float / deposit validation.", "r2-78-save.png"),
        ("Checker login", "MethmiB login OK via Use my password (PF-58560).", "chk-final.png"),
    ],
    "PF-58379": [
        ("Accrued Interest", "DONE N/A Kenya — feature not on cNwNb build.", "r2-na-79.png"),
    ],
    "PF-58380": [
        ("Receipt Reversal", "Reversal screen blank — no reverse rows.", "r2-80-rev.png"),
        ("Inquiry", "Seed account inquiry hit=false.", "r2-80-inq-004225.png"),
        ("Reallocation", "Reallocation screen empty.", "r2-80-realloc.png"),
    ],
    "PF-58381": [
        ("Document Request", "DONE N/A Kenya — not on Kenya COB/Lending build.", "r2-na-81.png"),
    ],
    "PF-58382": [
        ("Profit Sharing", "DONE N/A Kenya — Islamic profit-sharing not deployed.", "r2-na-82.png"),
    ],
    "PF-58383": [
        ("GBAF selector", "Deep routes trapped on Banking Type selector (PF-58398/58416).", "r2-83-account-management-manage-account.png"),
        ("Ownership transfer", "Ownership transfer blocked by selector.", "r2-83-maintenance-ownership-transfer.png"),
        ("Account inquiry", "Account inquiry blocked by selector.", "r2-83-cNwNb-account-inquiry.png"),
    ],
    "PF-58384": [
        ("BRWNS SMS", "DONE N/A Kenya — SMS/BRWNS not under Common Settings.", "r2-na-84.png"),
    ],
}


def find_src(name: str) -> Path | None:
    for root in (SRC, DONE, CHK):
        p = root / name
        if p.exists():
            return p
    return None


def annotate(name: str) -> Path | None:
    src = find_src(name)
    if not src:
        return None
    boxes = FINE.get(name, [(R(0.22, 0.15, 0.95, 0.85), f"ISSUE evidence: {name}")])
    im = Image.open(src).convert("RGBA")
    ov = Image.new("RGBA", im.size, (0, 0, 0, 0))
    d = ImageDraw.Draw(ov)
    for i, (xy, label) in enumerate(boxes):
        x1, y1, x2, y2 = xy
        d.rectangle([x1, y1, x2, y2], fill=(220, 20, 60, 45))
        for t in range(4):
            d.rectangle([x1 + t, y1 + t, x2 - t, y2 - t], outline=(200, 16, 46, 255))
        lab = f"{i+1}. {label}"[:95]
        tw = d.textlength(lab, font=FONT) if hasattr(d, "textlength") else len(lab) * 8
        ly = max(6, y1 - 24)
        d.rectangle([x1, ly, min(W - 8, x1 + tw + 12), ly + 22], fill=(200, 16, 46, 255))
        d.text((x1 + 6, ly + 3), lab, fill=(255, 255, 255, 255), font=FONT)
    out = Image.alpha_composite(im, ov).convert("RGB")
    d2 = ImageDraw.Draw(out)
    d2.rectangle([0, H - 26, W, H], fill=(25, 25, 25))
    d2.text((8, H - 22), f"FINE-TUNED | {name}", fill=(255, 210, 210), font=FONT_SM)
    ANN_OUT.mkdir(parents=True, exist_ok=True)
    dest = ANN_OUT / f"FT-{name}"
    out.save(dest, quality=92)
    return dest


def make_thumb(src: Path, dest: Path, max_w: int = 260) -> tuple[int, int]:
    im = Image.open(src).convert("RGB")
    w, h = im.size
    if w > max_w:
        nh = max(1, int(h * (max_w / w)))
        im = im.resize((max_w, nh), Image.Resampling.LANCZOS)
    else:
        max_w, nh = w, h
    if nh > 115:
        scale = 115 / nh
        max_w = max(1, int(max_w * scale))
        nh = 115
        im = Image.open(src).convert("RGB").resize((max_w, nh), Image.Resampling.LANCZOS)
    dest.parent.mkdir(parents=True, exist_ok=True)
    im.save(dest, format="PNG", optimize=True)
    return max_w, nh


def add_img(ws, row: int, path: Path, pw: int, ph: int) -> None:
    img = XLImage(str(path))
    img.width = pw
    img.height = ph
    marker = AnchorMarker(col=2, colOff=pixels_to_EMU(3), row=row - 1, rowOff=pixels_to_EMU(2))
    ext = XDRPositiveSize2D(pixels_to_EMU(pw), pixels_to_EMU(ph))
    img.anchor = OneCellAnchor(_from=marker, ext=ext)
    ws.add_image(img)


def write_story_excel(story: str, rows: list[tuple[str, str, str]], out_path: Path) -> dict:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    # Match iPay Lite Testing (1).xlsx column widths
    ws.column_dimensions["A"].width = 28.5
    ws.column_dimensions["B"].width = 133.0
    ws.column_dimensions["C"].width = 40.0

    placed = 0
    missing = []
    start = 2
    for i, (area, issue, src_name) in enumerate(rows):
        r = start + i
        ws.row_dimensions[r].height = 125
        a = ws.cell(r, 1, f"{story} | {area}")
        a.font = Font(bold=True, size=11)
        a.alignment = wrap
        a.border = thin
        b = ws.cell(r, 2, issue)
        b.alignment = wrap
        b.border = thin
        ws.cell(r, 3, "").border = thin

        ft = annotate(src_name)
        if not ft:
            ws.cell(r, 3, f"(missing {src_name})")
            missing.append(src_name)
            continue
        thumb = THUMB / f"{story}-{r}.png"
        pw, ph = make_thumb(ft, thumb)
        add_img(ws, r, thumb, pw, ph)
        placed += 1

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    # verify
    v = load_workbook(out_path)
    nimg = len(v["Sheet1"]._images)
    v.close()
    return {"story": story, "path": str(out_path), "rows": len(rows), "images": nimg, "missing": missing}


def main() -> None:
    for d in (OUT_DIR, DL_DIR, ART_DIR, ANN_OUT, THUMB):
        d.mkdir(parents=True, exist_ok=True)

    manifest = []
    for story, rows in STORY_ROWS.items():
        name = f"{story}-Kenya-UAT-Book1-visual-QA.xlsx"
        primary = OUT_DIR / name
        result = write_story_excel(story, rows, primary)
        for dest_dir in (DL_DIR, ART_DIR):
            shutil.copy2(primary, dest_dir / name)
        manifest.append(result)
        print(json.dumps(result))

    # index README
    idx = OUT_DIR / "00-INDEX.json"
    idx.write_text(json.dumps({"cycle": "2026-08-31 fine-tuned Book1 per-story", "files": manifest}, indent=2))
    (DL_DIR / "00-INDEX.json").write_text(idx.read_text())
    print("OUT_DIR", OUT_DIR)
    print("DL_DIR", DL_DIR)
    print("stories", len(manifest), "ann", len(list(ANN_OUT.glob("FT-*.png"))))


if __name__ == "__main__":
    main()
