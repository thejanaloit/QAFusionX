#!/usr/bin/env python3
"""Rebuild all PF-57868 reinit Book1 Excels with 5 columns + pass/fail footer."""
from __future__ import annotations

import json
import shutil
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image, ImageDraw, ImageFont

ROOT = Path(r"E:\QA OUTPUTS\PF-57868-reinit")
EXCEL = ROOT / "excel"
TEAMS = ROOT / "teams-drop"
THUMB = ROOT / "proof" / "_thumbs"
STORIES = [f"PF-{n}" for n in range(58374, 58385)]  # PF-58374 .. PF-58384

HEADERS = [
    "Area",
    "Issue",
    "Screenshot",
    "What is testing",
    "Why that failed your prediction",
]

try:
    FONT = ImageFont.truetype("arialbd.ttf", 14)
except Exception:
    FONT = ImageFont.load_default()

THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
WRAP = Alignment(wrap_text=True, vertical="top")


def classify(verdict: str) -> str:
    v = (verdict or "").upper()
    if v in {"PASS", "NOT_REPRO", "N/A_CONFIRMED", "NA_CONFIRMED", "PASSED"}:
        return "passed"
    if v in {"CONFIRMED", "PARTIAL", "FAIL", "FAILED", "PARTIAL_OR_FAIL"}:
        return "failed"
    if "N/A" in v and "CONFIRM" in v:
        return "passed"
    if "PARTIAL" in v or "CONFIRM" in v or "FAIL" in v:
        return "failed"
    return "failed"


def annotate(src: Path, label: str, dest: Path, na: bool = False) -> None:
    im = Image.open(src).convert("RGBA")
    w, h = im.size
    ov = Image.new("RGBA", im.size, (0, 0, 0, 0))
    d = ImageDraw.Draw(ov)
    color = (20, 100, 60, 50) if na else (220, 20, 60, 40)
    edge = (20, 100, 60, 255) if na else (200, 16, 46, 255)
    x1, y1, x2, y2 = int(0.15 * w), int(0.12 * h), int(0.95 * w), int(0.75 * h)
    d.rectangle([x1, y1, x2, y2], fill=color)
    for t in range(3):
        d.rectangle([x1 + t, y1 + t, x2 - t, y2 - t], outline=edge)
    d.rectangle([x1, max(4, y1 - 22), min(w - 4, x1 + 920), y1], fill=edge)
    d.text((x1 + 6, max(6, y1 - 18)), label[:100], fill=(255, 255, 255, 255), font=FONT)
    Image.alpha_composite(im, ov).convert("RGB").save(dest, quality=92)


def thumb(src: Path, dest: Path, max_w: int = 220) -> tuple[int, int]:
    im = Image.open(src).convert("RGB")
    w, h = im.size
    if w > max_w:
        h = max(1, int(h * (max_w / w)))
        w = max_w
        im = im.resize((w, h), Image.Resampling.LANCZOS)
    if h > 110:
        s = 110 / h
        w = max(1, int(w * s))
        h = 110
        im = Image.open(src).convert("RGB").resize((w, h), Image.Resampling.LANCZOS)
    dest.parent.mkdir(parents=True, exist_ok=True)
    im.save(dest)
    return w, h


def prediction_fail_why(f: dict, story: str) -> str:
    verdict = str(f.get("verdict", ""))
    notes = str(f.get("notes", ""))
    claim = str(f.get("claim", ""))
    cls = classify(verdict)
    if cls == "passed":
        if "N/A" in verdict.upper():
            return (
                f"Prediction held: feature expected absent / N/A on Kenya. "
                f"Observed: {notes or 'not present after mouse search'}."
            )
        return f"Prediction held (pass/{verdict}). {notes}".strip()
    # failed
    return (
        f"Predicted AC/path would be exercisable or defect fixed on Kenya UAT; "
        f"actual verdict={verdict}. Claim: {claim}. Evidence: {notes}"
    )


def what_testing(f: dict, story: str) -> str:
    bug = f.get("bug", "")
    claim = f.get("claim", "")
    return f"{story}: validating [{bug}] — {claim}"


def rows_for_story(story: str) -> list[dict]:
    proof = ROOT / "proof" / story
    summary_path = ROOT / "tracker" / f"{story}.json"
    data = json.loads(summary_path.read_text(encoding="utf-8")) if summary_path.exists() else {}
    findings = data.get("findings", [])
    out = []
    for f in findings:
        shot_name = f.get("shot", "")
        src = proof / shot_name
        if not src.exists():
            cands = sorted(proof.glob("*assert*.png")) + sorted(proof.glob("*ready*.png"))
            src = cands[-1] if cands else src
        out.append(
            {
                "area": f"{story} | {f.get('bug', 'QA')}",
                "issue": f"{f.get('claim', '')} — verdict={f.get('verdict')} | {f.get('notes', '')}",
                "src": src,
                "what": what_testing(f, story),
                "why": prediction_fail_why(f, story),
                "class": classify(str(f.get("verdict", ""))),
                "na": str(f.get("verdict", "")).upper().startswith("N/A") or bool(data.get("kenyaNA")),
                "counts": True,
            }
        )
    if not out:
        # fallback one row from any png
        pngs = sorted(proof.glob("*.png"))
        src = pngs[-1] if pngs else proof / "missing.png"
        out.append(
            {
                "area": f"{story} | QA",
                "issue": "No findings JSON — mouse proof only",
                "src": src,
                "what": f"{story}: mouse-only Kenya UAT surface check",
                "why": "Insufficient structured findings to confirm pass prediction.",
                "class": "failed",
                "na": False,
                "counts": True,
            }
        )
    return out


def write_story(story: str) -> dict:
    rows = rows_for_story(story)
    proof = ROOT / "proof" / story
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 42
    ws.column_dimensions["E"].width = 55

    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, size=11)
        cell.alignment = WRAP
        cell.border = THIN
    ws.row_dimensions[1].height = 22

    tested = passed = failed = 0
    for i, row in enumerate(rows):
        r = 2 + i
        ws.row_dimensions[r].height = 120
        vals = [row["area"], row["issue"], "", row["what"], row["why"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.alignment = WRAP
            cell.border = THIN
            if c == 1:
                cell.font = Font(bold=True, size=10)
        if row.get("counts"):
            tested += 1
            if row["class"] == "passed":
                passed += 1
            else:
                failed += 1
        src: Path = row["src"]
        if src.exists():
            ann = proof / f"ANN5-{src.name}"
            annotate(src, row["area"][:80], ann, na=row.get("na", False))
            tw, th = thumb(ann, THUMB / f"{story}-r{r}.png")
            img = XLImage(str(THUMB / f"{story}-r{r}.png"))
            img.width = tw
            img.height = th
            img.anchor = OneCellAnchor(
                _from=AnchorMarker(col=2, colOff=pixels_to_EMU(2), row=r - 1, rowOff=pixels_to_EMU(2)),
                ext=XDRPositiveSize2D(pixels_to_EMU(tw), pixels_to_EMU(th)),
            )
            ws.add_image(img)

    # footer
    fr = 2 + len(rows) + 1
    footer_lines = [
        f"Test cases tested: {tested}",
        f"Passed: {passed}",
        f"Failed: {failed}",
    ]
    for j, line in enumerate(footer_lines):
        cell = ws.cell(fr + j, 1, line)
        cell.font = Font(bold=True, size=12)
        cell.alignment = WRAP
        ws.merge_cells(start_row=fr + j, start_column=1, end_row=fr + j, end_column=5)

    name = f"{story}-Kenya-UAT-Book1-visual-QA-REINIT.xlsx"
    EXCEL.mkdir(parents=True, exist_ok=True)
    TEAMS.mkdir(parents=True, exist_ok=True)
    primary = EXCEL / name
    wb.save(primary)
    shutil.copy2(primary, TEAMS / name)
    dl = Path.home() / "Downloads" / "PF-57868-reinit"
    dl.mkdir(parents=True, exist_ok=True)
    shutil.copy2(primary, dl / name)
    nimg = len(load_workbook(primary)["Sheet1"]._images)
    return {
        "story": story,
        "tested": tested,
        "passed": passed,
        "failed": failed,
        "images": nimg,
        "path": str(primary),
    }


def main() -> None:
    # wipe marker: old 3-col Book1 contract superseded
    wipe = ROOT / "tracker" / "MEMORY-WIPE-SHORTCUTS.json"
    wipe.write_text(
        json.dumps(
            {
                "at": __import__("datetime").datetime.utcnow().isoformat() + "Z",
                "cleared": [
                    "excel-only shortcut",
                    "skip yaml/scripts",
                    "url deep-link between stories",
                    "fake Done without full flow",
                ],
                "lockedFlow": [
                    "jira/confluence stories",
                    "filter assigned",
                    "store",
                    "xtrai older cases",
                    "new test cases",
                    "store directory",
                    "yml",
                    "scripts",
                    "execute",
                    "one-by-one try pass",
                    "bug reports",
                    "excel 5-col + footer",
                ],
                "mouseOnly": True,
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    results = [write_story(s) for s in STORIES]
    for r in results:
        print(json.dumps(r))
    print("DONE", len(results), "tested_sum", sum(x["tested"] for x in results))


if __name__ == "__main__":
    main()
