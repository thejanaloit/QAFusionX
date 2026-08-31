#!/usr/bin/env python3
"""Write Book1-style Excel for one reinit story from fresh proof PNGs."""
from __future__ import annotations

import argparse
import json
import shutil
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image, ImageDraw, ImageFont

ROOT = Path(r"E:\QA OUTPUTS\PF-57868-reinit")
EXCEL = ROOT / "excel"
TEAMS = ROOT / "teams-drop"
THUMB = ROOT / "proof" / "_thumbs"

try:
    FONT = ImageFont.truetype("arialbd.ttf", 14)
    FONT_SM = ImageFont.truetype("arial.ttf", 12)
except Exception:
    FONT = ImageFont.load_default()
    FONT_SM = FONT


def annotate(src: Path, labels: list[str], dest: Path) -> Path:
    im = Image.open(src).convert("RGBA")
    w, h = im.size
    ov = Image.new("RGBA", im.size, (0, 0, 0, 0))
    d = ImageDraw.Draw(ov)
    # top banner + mid content box
    boxes = [
        (int(0.18 * w), int(0.12 * h), int(0.95 * w), int(0.28 * h)),
        (int(0.22 * w), int(0.30 * h), int(0.96 * w), int(0.72 * h)),
    ]
    for i, (x1, y1, x2, y2) in enumerate(boxes[: max(1, len(labels))]):
        d.rectangle([x1, y1, x2, y2], fill=(220, 20, 60, 40))
        for t in range(3):
            d.rectangle([x1 + t, y1 + t, x2 - t, y2 - t], outline=(200, 16, 46, 255))
        lab = (labels[i] if i < len(labels) else labels[-1])[:110]
        ly = max(4, y1 - 22)
        d.rectangle([x1, ly, min(w - 6, x1 + len(lab) * 8 + 16), ly + 20], fill=(200, 16, 46, 255))
        d.text((x1 + 5, ly + 2), lab, fill=(255, 255, 255, 255), font=FONT)
    out = Image.alpha_composite(im, ov).convert("RGB")
    d2 = ImageDraw.Draw(out)
    d2.rectangle([0, h - 24, w, h], fill=(20, 20, 20))
    d2.text((6, h - 20), f"REINIT | {src.name}", fill=(255, 210, 210), font=FONT_SM)
    dest.parent.mkdir(parents=True, exist_ok=True)
    out.save(dest, quality=92)
    return dest


def thumb(src: Path, dest: Path, max_w: int = 260) -> tuple[int, int]:
    im = Image.open(src).convert("RGB")
    w, h = im.size
    if w > max_w:
        nh = max(1, int(h * (max_w / w)))
        im = im.resize((max_w, nh), Image.Resampling.LANCZOS)
    else:
        max_w, nh = w, h
    if nh > 115:
        scale = 115 / nh
        max_w = max(1, int(max_w * scale))
        nh = 115
        im = Image.open(src).convert("RGB").resize((max_w, nh), Image.Resampling.LANCZOS)
    dest.parent.mkdir(parents=True, exist_ok=True)
    im.save(dest, format="PNG", optimize=True)
    return max_w, nh


def add_img(ws, row: int, path: Path, pw: int, ph: int) -> None:
    img = XLImage(str(path))
    img.width = pw
    img.height = ph
    marker = AnchorMarker(col=2, colOff=pixels_to_EMU(3), row=row - 1, rowOff=pixels_to_EMU(2))
    ext = XDRPositiveSize2D(pixels_to_EMU(pw), pixels_to_EMU(ph))
    img.anchor = OneCellAnchor(_from=marker, ext=ext)
    ws.add_image(img)


def write_excel(story: str, rows: list[dict], out_path: Path) -> dict:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    ws.column_dimensions["A"].width = 28.5
    ws.column_dimensions["B"].width = 133.0
    ws.column_dimensions["C"].width = 40.0

    placed = 0
    missing = []
    for i, row in enumerate(rows):
        r = 2 + i
        ws.row_dimensions[r].height = 125
        a = ws.cell(r, 1, f"{story} | {row['area']}")
        a.font = Font(bold=True, size=11)
        a.alignment = wrap
        a.border = thin
        b = ws.cell(r, 2, row["issue"])
        b.alignment = wrap
        b.border = thin
        ws.cell(r, 3, "").border = thin

        src = Path(row["shot"])
        if not src.exists():
            ws.cell(r, 3, f"(missing {src.name})")
            missing.append(src.name)
            continue
        ann = ROOT / "proof" / story / f"ANN-{src.name}"
        annotate(src, [row.get("label", row["issue"][:90])], ann)
        tpath = THUMB / f"{story}-{r}.png"
        pw, ph = thumb(ann, tpath)
        add_img(ws, r, tpath, pw, ph)
        placed += 1

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    v = load_workbook(out_path)
    nimg = len(v["Sheet1"]._images)
    v.close()
    return {"story": story, "path": str(out_path), "rows": len(rows), "images": nimg, "missing": missing}


def from_summary(story: str) -> list[dict]:
    summary_path = ROOT / "tracker" / f"{story}.json"
    proof = ROOT / "proof" / story
    rows: list[dict] = []
    if summary_path.exists():
        data = json.loads(summary_path.read_text(encoding="utf-8"))
        for f in data.get("findings", []):
            shot = proof / f.get("shot", "")
            if not shot.exists():
                # find assert shots
                cands = sorted(proof.glob("*assert*.png")) + sorted(proof.glob("*ready*.png"))
                shot = cands[-1] if cands else shot
            rows.append(
                {
                    "area": f.get("bug", "QA"),
                    "issue": f"{f.get('claim', '')} — verdict={f.get('verdict')} | {f.get('notes', '')}",
                    "shot": str(shot),
                    "label": f"{f.get('bug')}: {f.get('verdict')}",
                }
            )
    if not rows:
        shots = sorted(proof.glob("*.png"))[-4:]
        for s in shots:
            rows.append(
                {
                    "area": "Reinit proof",
                    "issue": f"{story} reinit mouse-only evidence: {s.name}",
                    "shot": str(s),
                    "label": s.name,
                }
            )
    return rows


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--story", required=True)
    args = ap.parse_args()
    story = args.story
    rows = from_summary(story)
    name = f"{story}-Kenya-UAT-Book1-visual-QA-REINIT.xlsx"
    primary = EXCEL / name
    result = write_excel(story, rows, primary)
    TEAMS.mkdir(parents=True, exist_ok=True)
    shutil.copy2(primary, TEAMS / name)
    dl = Path.home() / "Downloads" / "PF-57868-reinit"
    dl.mkdir(parents=True, exist_ok=True)
    shutil.copy2(primary, dl / name)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
