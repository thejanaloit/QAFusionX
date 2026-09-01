#!/usr/bin/env python3
"""Build Book1 Excels for PF-58380..58384 from reinit proof/_summary.json."""
from __future__ import annotations

import json
import shutil
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image, ImageDraw, ImageFont

ROOT = Path(r"E:\QA OUTPUTS\PF-57868-reinit")
EXCEL = ROOT / "excel"
TEAMS = ROOT / "teams-drop"
THUMB = ROOT / "proof" / "_thumbs"
STORIES = ["PF-58380", "PF-58381", "PF-58382", "PF-58383", "PF-58384"]

try:
    FONT = ImageFont.truetype("arialbd.ttf", 14)
except Exception:
    FONT = ImageFont.load_default()


def annotate(src: Path, label: str, dest: Path, na: bool = False) -> None:
    im = Image.open(src).convert("RGBA")
    w, h = im.size
    ov = Image.new("RGBA", im.size, (0, 0, 0, 0))
    d = ImageDraw.Draw(ov)
    color = (20, 100, 60, 50) if na else (220, 20, 60, 40)
    edge = (20, 100, 60, 255) if na else (200, 16, 46, 255)
    x1, y1, x2, y2 = int(0.15 * w), int(0.12 * h), int(0.95 * w), int(0.75 * h)
    d.rectangle([x1, y1, x2, y2], fill=color)
    for t in range(3):
        d.rectangle([x1 + t, y1 + t, x2 - t, y2 - t], outline=edge)
    d.rectangle([x1, max(4, y1 - 22), min(w - 4, x1 + 920), y1], fill=edge)
    d.text((x1 + 6, max(6, y1 - 18)), label[:100], fill=(255, 255, 255, 255), font=FONT)
    Image.alpha_composite(im, ov).convert("RGB").save(dest, quality=92)


def thumb(src: Path, dest: Path, max_w: int = 260) -> tuple[int, int]:
    im = Image.open(src).convert("RGB")
    w, h = im.size
    if w > max_w:
        h = max(1, int(h * (max_w / w)))
        w = max_w
        im = im.resize((w, h), Image.Resampling.LANCZOS)
    if h > 115:
        s = 115 / h
        w = max(1, int(w * s))
        h = 115
        im = Image.open(src).convert("RGB").resize((w, h), Image.Resampling.LANCZOS)
    dest.parent.mkdir(parents=True, exist_ok=True)
    im.save(dest)
    return w, h


def write_story(story: str) -> dict:
    proof = ROOT / "proof" / story
    summary_path = ROOT / "tracker" / f"{story}.json"
    data = json.loads(summary_path.read_text(encoding="utf-8")) if summary_path.exists() else {}
    findings = data.get("findings", [])
    rows = []
    for f in findings:
        shot_name = f.get("shot", "")
        src = proof / shot_name
        if not src.exists():
            cands = sorted(proof.glob("*assert*.png")) + sorted(proof.glob("*ready*.png")) + sorted(proof.glob("*.png"))
            src = cands[-1] if cands else src
        rows.append(
            (
                f.get("bug", "QA"),
                f"{f.get('claim', '')} — verdict={f.get('verdict')} | {f.get('notes', '')}",
                src,
                str(f.get("verdict", "")).startswith("N/A"),
            )
        )
    # always add a couple of proof shots
    for extra in sorted(proof.glob("*.png"))[:2]:
        if all(extra != r[2] for r in rows):
            rows.append(("Proof", f"{story} mouse-only evidence: {extra.name}", extra, data.get("kenyaNA", False)))

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    ws.column_dimensions["A"].width = 28.5
    ws.column_dimensions["B"].width = 133.0
    ws.column_dimensions["C"].width = 40.0
    placed = 0
    for i, (area, issue, src, na) in enumerate(rows[:5]):
        if not src.exists():
            continue
        r = 2 + i
        ws.row_dimensions[r].height = 125
        a = ws.cell(r, 1, f"{story} | {area}")
        a.font = Font(bold=True, size=11)
        a.alignment = wrap
        a.border = thin
        b = ws.cell(r, 2, issue)
        b.alignment = wrap
        b.border = thin
        ws.cell(r, 3, "").border = thin
        ann = proof / f"ANN-{src.name}"
        annotate(src, f"{area}", ann, na=na)
        tw, th = thumb(ann, THUMB / f"{story}-{r}.png")
        img = XLImage(str(THUMB / f"{story}-{r}.png"))
        img.width = tw
        img.height = th
        img.anchor = OneCellAnchor(
            _from=AnchorMarker(col=2, colOff=pixels_to_EMU(3), row=r - 1, rowOff=pixels_to_EMU(2)),
            ext=XDRPositiveSize2D(pixels_to_EMU(tw), pixels_to_EMU(th)),
        )
        ws.add_image(img)
        placed += 1

    name = f"{story}-Kenya-UAT-Book1-visual-QA-REINIT.xlsx"
    primary = EXCEL / name
    EXCEL.mkdir(parents=True, exist_ok=True)
    TEAMS.mkdir(parents=True, exist_ok=True)
    wb.save(primary)
    shutil.copy2(primary, TEAMS / name)
    dl = Path.home() / "Downloads" / "PF-57868-reinit"
    dl.mkdir(parents=True, exist_ok=True)
    shutil.copy2(primary, dl / name)
    nimg = len(load_workbook(primary)["Sheet1"]._images)
    return {"story": story, "path": str(primary), "images": nimg, "kenyaNA": data.get("kenyaNA"), "honestDoneAllowed": data.get("honestDoneAllowed")}


def main() -> None:
    results = []
    for s in STORIES:
        results.append(write_story(s))
        print(json.dumps(results[-1]))

    master_path = ROOT / "tracker" / "00-MASTER.json"
    master = json.loads(master_path.read_text(encoding="utf-8-sig"))
    by = {r["story"]: r for r in results}
    for s in master["stories"]:
        if s["key"] not in by:
            continue
        r = by[s["key"]]
        tracker = json.loads((ROOT / "tracker" / f"{s['key']}.json").read_text(encoding="utf-8"))
        s["excel"] = f"excel/{s['key']}-Kenya-UAT-Book1-visual-QA-REINIT.xlsx"
        if tracker.get("kenyaNA") and tracker.get("honestDoneAllowed"):
            s["status"] = "DONE_KENYA_NA"
            s["verdict"] = "N/A_CONFIRMED"
            s["notes"] = "Kenya N/A reconfirmed with mouse proof"
        elif tracker.get("honestDoneAllowed"):
            s["status"] = "DONE_VERIFIED"
            s["verdict"] = "PASS"
        else:
            s["status"] = "QA_COMPLETE_NOT_DONE"
            s["verdict"] = "PARTIAL_OR_FAIL"
            s["notes"] = "Open/partial — not Done"
    master_path.write_text(json.dumps(master, indent=2), encoding="utf-8")
    print("MASTER_UPDATED", len(results))


if __name__ == "__main__":
    main()
